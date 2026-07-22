"""Semantic Registry management API routes."""
from __future__ import annotations

import asyncio
import json
import logging
import time
import uuid
from datetime import datetime

from fastapi import APIRouter, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/api/v1/medical-insurance-ai-agent/semantic",
    tags=["semantic-registry"],
)


class ObjectSummary(BaseModel):
    object_code: str
    name: str
    domain_code: str
    status: str
    current_version: str | None = None
    definition: str | None = None
    identifier: str | None = None
    version: str | None = None


class ObjectDetail(BaseModel):
    object_code: str
    name: str
    definition: str | None
    domain_code: str
    identifier: str | None
    source_object: str | None
    source_adapter_port: str | None
    relations: list[dict]
    version: str
    status: str
    current_version: str | None = None
    locked_by_skills: list[dict] = []


class MetricSummary(BaseModel):
    metric_code: str
    name: str
    object_code: str
    metric_type: str
    importance: str
    status: str
    usage_count: int = 0


class MetricDetail(BaseModel):
    metric_code: str
    name: str
    definition: str | None
    object_code: str
    metric_type: str
    semantic_type: str | None
    unit: str | None
    required: bool
    importance: str
    value_domain: str | None
    source_object: str | None
    source_field: str | None
    source_adapter_port: str | None
    usage_count: int
    quality_score: float
    version: str
    status: str


class DomainProgress(BaseModel):
    domain_code: str
    name: str
    total_metrics: int
    mapped_metrics: int
    percentage: float
    skill_refs: int = 0


class SemanticSummary(BaseModel):
    domains_count: int
    objects_count: int
    metrics_count: int
    mapped_count: int
    unmapped_count: int
    value_missing_count: int = 0
    mapping_rate: float
    skill_references: int
    domain_progress: list[DomainProgress]
    discovery_tables: int = 0
    discovery_fields: int = 0
    discovery_unmapped: int = 0


class FieldMetadataItem(BaseModel):
    field_name: str
    description: str
    data_type: str
    non_null_rate: float
    distinct_count: int
    last_updated: str
    sample_value: str | None


class FieldMetadataResponse(BaseModel):
    table: str
    adapter: str
    fields: list[FieldMetadataItem]


class ValueMismatchItem(BaseModel):
    value: str
    count: int
    percentage: float
    mapped_to: str | None
    status: str  # "mapped" | "unmapped"


class ValueMismatchResponse(BaseModel):
    metric_code: str
    value_domain: str | None
    source_values: list[ValueMismatchItem]


class ValueMappingRequest(BaseModel):
    domain_code: str
    source_value: str
    standard_value: str


class ValueMappingResponse(BaseModel):
    status: str
    domain_code: str
    source_value: str
    standard_value: str


class DiscoveryScanResponse(BaseModel):
    task_id: str
    status: str


class DiscoverySampleStats(BaseModel):
    """样本值统计信息。按字段类型区分：时间/数值 → max/min/top_freq；字符串 → enum_type。"""
    max: str | None = None
    min: str | None = None
    top_freq: list[dict] | None = None  # [{"value": "...", "count": 123}]
    enum_type: str | None = None  # "枚举类型" | "海量枚举类型"
    is_long_text: bool = False
    non_null_count: int = 0


class ValueScoreDetail(BaseModel):
    total: float = 0.0
    grade: str = "E"
    non_null_score: float = 0.0
    non_null_rate: float = 0.0
    desc_score: int = 0
    has_desc: bool = False
    sample_score: int = 0
    has_sample: bool = False
    recency_score: int = 0
    last_updated: str | None = None
    usage_score: int = 0
    usage_count: int = 0


class DiscoveryResultItem(BaseModel):
    field_name: str
    table_name: str
    description: str | None
    data_type: str
    non_null_rate: float
    non_null_row_count: int = 0
    distinct_count: int | None = None
    sample_value: str | None
    sample_values: list[str] | None = None
    sample_stats: DiscoverySampleStats | None = None
    is_dictionary: bool = False
    last_updated: str | None = None
    suggested_object: str | None
    mapped: bool
    table_schema: str | None = None
    is_nullable: str | None = None
    is_primary_key: bool = False
    remark: str | None = None
    quality_score: float = 0.0
    value_score: ValueScoreDetail | None = None


class DiscoveryResultsResponse(BaseModel):
    tables_count: int
    fields_count: int
    mapped_fields: int
    unmapped_fields: int
    fields: list[DiscoveryResultItem]


class DiscoveryHistoryItem(BaseModel):
    scan_id: str
    started_at: str | None = None
    duration_seconds: float | None = None
    status: str
    tables_scanned: int
    unmapped_found: int
    new_found: int

# Mock field metadata for known tables — real SQL statistics will replace these later.
_MOCK_FIELD_METADATA: dict[str, dict[str, list[dict]]] = {
    "InsuranceInterfacePort": {
        "yb_settlement": [
            {"field_name": "SET_NO", "description": "结算流水号", "data_type": "varchar(32)", "non_null_rate": 1.0, "distinct_count": 48512, "last_updated": "2026-07-14T23:59:00", "sample_value": "S202607140001"},
            {"field_name": "INSU_TYPE", "description": "医保类型编码", "data_type": "varchar(10)", "non_null_rate": 1.0, "distinct_count": 6, "last_updated": "2026-07-14T23:59:00", "sample_value": "310"},
            {"field_name": "HOSP_LV", "description": "医院等级", "data_type": "varchar(10)", "non_null_rate": 1.0, "distinct_count": 3, "last_updated": "2026-07-14T23:59:00", "sample_value": "3"},
            {"field_name": "ADM_DATE", "description": "入院日期", "data_type": "datetime", "non_null_rate": 1.0, "distinct_count": 732, "last_updated": "2026-07-14T23:59:00", "sample_value": "2026-07-01"},
            {"field_name": "DSCG_DATE", "description": "出院日期", "data_type": "datetime", "non_null_rate": 0.97, "distinct_count": 728, "last_updated": "2026-07-14T23:59:00", "sample_value": "2026-07-10"},
            {"field_name": "TOTAL_FEE", "description": "总费用（元）", "data_type": "decimal(18,2)", "non_null_rate": 1.0, "distinct_count": 32104, "last_updated": "2026-07-14T23:59:00", "sample_value": "12586.40"},
            {"field_name": "FUND_PAY", "description": "统筹基金支付（元）", "data_type": "decimal(18,2)", "non_null_rate": 0.94, "distinct_count": 18723, "last_updated": "2026-07-14T23:59:00", "sample_value": "8234.50"},
            {"field_name": "QFX", "description": "起付线标准（元）", "data_type": "decimal(18,2)", "non_null_rate": 0.88, "distinct_count": 15, "last_updated": "2026-07-14T23:59:00", "sample_value": "800.00"},
            {"field_name": "QFY", "description": "起付线内金额（元）—— 已废弃，仅历史数据保留", "data_type": "decimal(18,2)", "non_null_rate": 0.03, "distinct_count": 412, "last_updated": "2025-06-30T23:59:00", "sample_value": "null"},
            {"field_name": "SELF_PAY", "description": "个人自付金额（元）", "data_type": "decimal(18,2)", "non_null_rate": 0.98, "distinct_count": 24105, "last_updated": "2026-07-14T23:59:00", "sample_value": "3120.80"},
            {"field_name": "MED_FEE", "description": "医保范围内费用（元）", "data_type": "decimal(18,2)", "non_null_rate": 0.96, "distinct_count": 28901, "last_updated": "2026-07-14T23:59:00", "sample_value": "11230.00"},
            {"field_name": "CREATE_TIME", "description": "创建时间", "data_type": "datetime", "non_null_rate": 1.0, "distinct_count": 48210, "last_updated": "2026-07-14T23:59:00", "sample_value": "2026-07-14 10:30:00"},
        ],
        "yb_fee_detail": [
            {"field_name": "DETAIL_NO", "description": "费用明细流水号", "data_type": "varchar(32)", "non_null_rate": 1.0, "distinct_count": 1523400, "last_updated": "2026-07-14T23:59:00", "sample_value": "FD2026071400001"},
            {"field_name": "SET_NO", "description": "结算流水号（关联 yb_settlement）", "data_type": "varchar(32)", "non_null_rate": 1.0, "distinct_count": 48512, "last_updated": "2026-07-14T23:59:00", "sample_value": "S202607140001"},
            {"field_name": "ITEM_CODE", "description": "收费项目编码", "data_type": "varchar(20)", "non_null_rate": 1.0, "distinct_count": 3200, "last_updated": "2026-07-14T23:59:00", "sample_value": "110100001"},
            {"field_name": "ITEM_NAME", "description": "收费项目名称", "data_type": "varchar(100)", "non_null_rate": 1.0, "distinct_count": 2850, "last_updated": "2026-07-14T23:59:00", "sample_value": "普通挂号费"},
            {"field_name": "QTY", "description": "数量", "data_type": "decimal(12,2)", "non_null_rate": 1.0, "distinct_count": 120, "last_updated": "2026-07-14T23:59:00", "sample_value": "1.00"},
            {"field_name": "PRICE", "description": "单价（元）", "data_type": "decimal(12,2)", "non_null_rate": 0.99, "distinct_count": 850, "last_updated": "2026-07-14T23:59:00", "sample_value": "20.00"},
            {"field_name": "TOTAL_AMOUNT", "description": "金额小计（元）", "data_type": "decimal(18,2)", "non_null_rate": 1.0, "distinct_count": 48200, "last_updated": "2026-07-14T23:59:00", "sample_value": "20.00"},
            {"field_name": "SELF_PAY_AMOUNT", "description": "自付金额（元）", "data_type": "decimal(18,2)", "non_null_rate": 0.85, "distinct_count": 32100, "last_updated": "2026-07-14T23:59:00", "sample_value": "0.00"},
            {"field_name": "FUND_PAY_AMOUNT", "description": "统筹支付金额（元）", "data_type": "decimal(18,2)", "non_null_rate": 0.78, "distinct_count": 29800, "last_updated": "2026-07-14T23:59:00", "sample_value": "15.00"},
        ],
    },
}


# Mock value distributions for metrics with value_domain set.
_MOCK_VALUE_MISMATCHES: dict[str, list[dict]] = {
    "Settlement.hospital_level": [
        {"value": "LEVEL_3", "count": 42300, "percentage": 68.0, "mapped_to": "三级", "status": "mapped"},
        {"value": "LEVEL_2", "count": 15800, "percentage": 25.4, "mapped_to": "二级", "status": "mapped"},
        {"value": "LEVEL_1", "count": 3200, "percentage": 5.1, "mapped_to": "一级", "status": "mapped"},
        {"value": "3A", "count": 850, "percentage": 1.4, "mapped_to": None, "status": "unmapped"},
        {"value": "null", "count": 620, "percentage": 1.0, "mapped_to": None, "status": "unmapped"},
    ],
    "Settlement.person_type": [
        {"value": "EMPLOYED", "count": 30000, "percentage": 60.0, "mapped_to": "在职", "status": "mapped"},
        {"value": "RETIRED", "count": 18000, "percentage": 36.0, "mapped_to": "退休", "status": "mapped"},
        {"value": "城居", "count": 2000, "percentage": 4.0, "mapped_to": None, "status": "unmapped"},
    ],
    "Settlement.insurance_type": [
        {"value": "EMPLOYEE", "count": 35000, "percentage": 70.0, "mapped_to": "城镇职工", "status": "mapped"},
        {"value": "RESIDENT", "count": 14000, "percentage": 28.0, "mapped_to": "城乡居民", "status": "mapped"},
        {"value": "31", "count": 800, "percentage": 1.6, "mapped_to": None, "status": "unmapped"},
        {"value": "39", "count": 200, "percentage": 0.4, "mapped_to": None, "status": "unmapped"},
    ],
}


@router.get("/field-metadata", response_model=FieldMetadataResponse)
def get_field_metadata(
    adapter: str = Query(..., description="Adapter port name"),
    table: str = Query(..., description="Table name"),
):
    """Return field quality metadata for a given adapter port and table.

    MVP uses hardcoded mock data. Real SQL statistics will be wired later.
    """
    adapter_data = _MOCK_FIELD_METADATA.get(adapter, {})
    fields_data = adapter_data.get(table, [])
    return FieldMetadataResponse(
        table=table,
        adapter=adapter,
        fields=[FieldMetadataItem(**f) for f in fields_data],
    )


def get_registry():
    """委托给语义层的全局单例（集中维护，避免服务层反向依赖路由层）。"""
    from src.semantic_layer.registry import get_semantic_registry
    return get_semantic_registry()


class CreateObjectRequest(BaseModel):
    object_code: str
    domain_code: str
    name: str
    definition: str | None = None


class UpdateObjectRequest(BaseModel):
    name: str | None = None
    definition: str | None = None
    domain_code: str | None = None
    identifier: str | None = None
    source_object: str | None = None
    source_adapter_port: str | None = None
    version: str | None = None
    status: str | None = None


@router.post("/objects")
def create_object(req: CreateObjectRequest):
    """创建新的业务对象。"""
    from src.semantic_layer.models import BusinessObject
    reg = get_registry()
    store = reg._store
    if store.get_object(req.object_code):
        raise HTTPException(status_code=409, detail=f"对象 '{req.object_code}' 已存在")
    if not store.get_domain(req.domain_code):
        raise HTTPException(status_code=400, detail=f"域 '{req.domain_code}' 不存在，请先创建域")
    obj = BusinessObject(object_code=req.object_code, domain_code=req.domain_code, name=req.name, definition=req.definition)
    store.save_object(obj)
    return {"status": "ok", "object_code": req.object_code, "name": req.name}


@router.put("/objects/{object_code}")
def update_object(object_code: str, req: UpdateObjectRequest):
    """更新业务对象（名称、描述、域、数据源、版本、状态等）。"""
    reg = get_registry()
    store = reg._store
    obj = store.get_object(object_code)
    if not obj:
        raise HTTPException(status_code=404, detail=f"对象 '{object_code}' 不存在")
    if req.name is not None:
        obj.name = req.name
    if req.definition is not None:
        obj.definition = req.definition if req.definition.strip() else None
    if req.domain_code is not None:
        if not store.get_domain(req.domain_code):
            raise HTTPException(status_code=400, detail=f"域 '{req.domain_code}' 不存在")
        obj.domain_code = req.domain_code
    if req.identifier is not None:
        obj.identifier = req.identifier
    if req.source_object is not None:
        obj.source_object = req.source_object
    if req.source_adapter_port is not None:
        obj.source_adapter_port = req.source_adapter_port
    if req.version is not None:
        obj.version = req.version
    if req.status is not None:
        obj.status = req.status
    store.save_object(obj)
    return {"status": "ok", "object_code": object_code, "name": obj.name, "version": obj.version}


@router.delete("/objects/{object_code}")
def delete_object(object_code: str):
    """删除业务对象。"""
    reg = get_registry()
    store = reg._store
    obj = store.get_object(object_code)
    if not obj:
        raise HTTPException(status_code=404, detail=f"对象 '{object_code}' 不存在")
    store.delete_object(object_code)
    return {"status": "ok", "object_code": object_code}


@router.get("/objects", response_model=list[ObjectSummary])
def list_objects(domain_code: str | None = Query(None)):
    reg = get_registry()
    objects = reg.list_objects(domain_code)
    return [ObjectSummary(object_code=o.object_code, name=o.name, domain_code=o.domain_code, status=o.status, current_version=o.current_version, definition=o.definition, identifier=o.identifier, version=o.version) for o in objects]


@router.get("/objects/{object_code}", response_model=ObjectDetail)
def get_object(object_code: str):
    reg = get_registry()
    obj = reg.get_object(object_code)
    if obj is None:
        raise HTTPException(status_code=404, detail=f"Object '{object_code}' not found")
    return ObjectDetail(
        object_code=obj.object_code, name=obj.name, definition=obj.definition,
        domain_code=obj.domain_code, identifier=obj.identifier,
        source_object=obj.source_object, source_adapter_port=obj.source_adapter_port,
        relations=[r.model_dump() for r in obj.relations], version=obj.version, status=obj.status,
        current_version=obj.current_version,
        locked_by_skills=_compute_skill_locks().get(object_code, []),
    )


# ── 对象版本快照（阶段2）──
class VersionMetricInfo(BaseModel):
    metric_code: str
    name: str
    semantic_type: str | None = None
    required: bool
    source_field: str | None = None
    importance: str


class ObjectVersionInfo(BaseModel):
    version_id: str
    object_code: str
    version: str
    published_at: datetime
    published_by: str | None = None
    changelog: str | None = None
    metric_count: int


class ObjectVersionDetail(ObjectVersionInfo):
    snapshot: dict
    metrics: list[VersionMetricInfo]


class PublishObjectRequest(BaseModel):
    changelog: str | None = None
    published_by: str | None = None


@router.post("/objects/{object_code}/publish", response_model=ObjectVersionInfo)
def publish_object(object_code: str, req: PublishObjectRequest):
    """发布对象：冻结当前草稿指标为不可变版本快照。"""
    reg = get_registry()
    try:
        version = reg.publish_object(
            object_code, changelog=req.changelog, published_by=req.published_by)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return ObjectVersionInfo(
        version_id=version.version_id, object_code=version.object_code,
        version=version.version, published_at=version.published_at,
        published_by=version.published_by, changelog=version.changelog,
        metric_count=len(version.metrics))


@router.get("/objects/{object_code}/versions", response_model=list[ObjectVersionInfo])
def list_object_versions(object_code: str):
    """列出对象所有发布版本。"""
    reg = get_registry()
    versions = reg.list_object_versions(object_code)
    return [ObjectVersionInfo(
        version_id=v.version_id, object_code=v.object_code, version=v.version,
        published_at=v.published_at, published_by=v.published_by,
        changelog=v.changelog, metric_count=len(v.metrics)) for v in versions]


@router.get("/objects/{object_code}/versions/{version}", response_model=ObjectVersionDetail)
def get_object_version(object_code: str, version: str):
    """查看指定版本快照详情（含冻结的指标列表）。"""
    reg = get_registry()
    v = reg.get_object_version(object_code, version)
    if v is None:
        raise HTTPException(status_code=404, detail=f"版本 {object_code}@{version} 不存在")
    return ObjectVersionDetail(
        version_id=v.version_id, object_code=v.object_code, version=v.version,
        published_at=v.published_at, published_by=v.published_by, changelog=v.changelog,
        metric_count=len(v.metrics), snapshot=v.snapshot,
        metrics=[VersionMetricInfo(
            metric_code=m.metric_code, name=m.name, semantic_type=m.semantic_type,
            required=m.required, source_field=m.source_field, importance=m.importance
        ) for m in v.metrics])


@router.get("/metrics", response_model=list[MetricSummary])
def list_metrics(object_code: str | None = Query(None)):
    reg = get_registry()
    metrics = reg.get_metrics_by_object(object_code) if object_code else []
    refs = _get_skill_metric_refs()
    return [MetricSummary(metric_code=m.metric_code, name=m.name, object_code=m.object_code, metric_type=m.metric_type, importance=m.importance, status=m.status, usage_count=refs.get(m.metric_code, 0)) for m in metrics]


@router.get("/metrics/{metric_code:path}/value-mismatch", response_model=ValueMismatchResponse)
def get_value_mismatch(metric_code: str):
    """Return all distinct values for a metric's source field, with mapping status."""
    reg = get_registry()
    metric = reg.get_metric(metric_code)
    if metric is None:
        raise HTTPException(status_code=404, detail=f"Metric '{metric_code}' not found")
    source_values_data = _MOCK_VALUE_MISMATCHES.get(metric_code, [])
    return ValueMismatchResponse(
        metric_code=metric_code,
        value_domain=metric.value_domain,
        source_values=[ValueMismatchItem(**v) for v in source_values_data],
    )


@router.get("/metrics/{metric_code:path}", response_model=MetricDetail)
def get_metric(metric_code: str):
    reg = get_registry()
    metric = reg.get_metric(metric_code)
    if metric is None:
        raise HTTPException(status_code=404, detail=f"Metric '{metric_code}' not found")
    # GET 只读：质量分直接返回已持久化的值（在保存映射或扫描完成时计算写入）
    # usage_count 取静态技能引用数（物理编码与语义编码不一致，运行时累加不可靠）
    usage_count = _get_skill_metric_refs().get(metric.metric_code, metric.usage_count)
    return MetricDetail(
        metric_code=metric.metric_code, name=metric.name, definition=metric.definition,
        object_code=metric.object_code, metric_type=metric.metric_type,
        semantic_type=metric.semantic_type, unit=metric.unit, required=metric.required,
        importance=metric.importance, value_domain=metric.value_domain,
        source_object=metric.source_object, source_field=metric.source_field,
        source_adapter_port=metric.source_adapter_port,
        usage_count=usage_count, quality_score=metric.quality_score,
        version=metric.version, status=metric.status,
    )


class BatchCreateMetricItem(BaseModel):
    object_code: str
    name: str
    metric_code: str | None = None  # 如果指定则直接使用，否则自动生成 {object_code}.{name}
    metric_type: str = "Atomic"
    semantic_type: str = "Amount"
    definition: str | None = None
    unit: str | None = None
    importance: str = "optional"
    value_domain: str | None = None
    required: bool = False
    source_table: str | None = None
    source_field: str | None = None


class BatchCreateMetricsRequest(BaseModel):
    items: list[BatchCreateMetricItem]


class BatchCreateMetricResult(BaseModel):
    index: int
    metric_code: str
    name: str
    status: str  # "created" | "skipped" | "error"
    error: str | None = None


@router.post("/metrics")
def create_metric(req: CreateMetricRequest):
    """创建新的业务指标。metric_code 自动生成为 {object_code}.{name_pascal}。"""
    from src.semantic_layer.models import Metric
    reg = get_registry()
    store = reg._store

    # 校验对象存在
    obj = store.get_object(req.object_code)
    if not obj:
        raise HTTPException(status_code=400, detail=f"对象 '{req.object_code}' 不存在，请先创建对象")

    # 自动生成 metric_code
    name_pascal = req.name.strip().replace(" ", "_")
    metric_code = f"{req.object_code}.{name_pascal}"

    if store.get_metric(metric_code):
        raise HTTPException(status_code=409, detail=f"指标 '{metric_code}' 已存在")

    metric = Metric(
        metric_code=metric_code,
        object_code=req.object_code,
        name=req.name.strip(),
        definition=req.definition,
        metric_type=req.metric_type,
        semantic_type=req.semantic_type,
        unit=req.unit,
        importance=req.importance,
        value_domain=req.value_domain,
        required=req.required,
    )
    if req.source_field:
        metric.source_field = req.source_field
    if req.source_table:
        metric.source_object = req.source_table
    store.save_metric(metric)
    return {"status": "ok", "metric_code": metric_code, "name": metric.name}


@router.post("/metrics/batch", response_model=list[BatchCreateMetricResult])
def create_metrics_batch(req: BatchCreateMetricsRequest):
    """批量创建业务指标。返回每个 item 的结果（创建/跳过/错误）。"""
    from src.semantic_layer.models import Metric
    reg = get_registry()
    store = reg._store

    results: list[BatchCreateMetricResult] = []
    for i, item in enumerate(req.items):
        metric_code = item.metric_code if item.metric_code and '.' in item.metric_code else f"{item.object_code}.{item.metric_code or item.name.strip().replace(' ', '_')}"
        try:
            obj = store.get_object(item.object_code)
            if not obj:
                results.append(BatchCreateMetricResult(
                    index=i, metric_code=metric_code, name=item.name, status="error",
                    error=f"对象 '{item.object_code}' 不存在",
                ))
                continue

            if store.get_metric(metric_code):
                results.append(BatchCreateMetricResult(
                    index=i, metric_code=metric_code, name=item.name, status="skipped",
                    error="指标已存在",
                ))
                continue

            metric = Metric(
                metric_code=metric_code,
                object_code=item.object_code,
                name=item.name.strip(),
                definition=item.definition,
                metric_type=item.metric_type,
                semantic_type=item.semantic_type,
                unit=item.unit,
                importance=item.importance,
                value_domain=item.value_domain,
                required=item.required,
            )
            if item.source_field:
                metric.source_field = item.source_field
            if item.source_table:
                metric.source_object = item.source_table
            store.save_metric(metric)
            results.append(BatchCreateMetricResult(
                index=i, metric_code=metric_code, name=item.name, status="created",
            ))
        except Exception as exc:
            logger.exception("批量创建指标失败 index=%d", i)
            results.append(BatchCreateMetricResult(
                index=i, metric_code=metric_code, name=item.name, status="error",
                error=str(exc),
            ))

    return results


@router.put("/metrics/{metric_code:path}")
def update_metric(metric_code: str, req: UpdateMetricRequest):
    """更新业务指标（名称、描述、类型、编码、映射等）。"""
    reg = get_registry()
    store = reg._store
    metric = store.get_metric(metric_code)
    if not metric:
        raise HTTPException(status_code=404, detail=f"指标 '{metric_code}' 不存在")

    # Handle metric_code rename: delete old, insert new
    if req.metric_code is not None and req.metric_code != metric_code:
        if store.get_metric(req.metric_code):
            raise HTTPException(status_code=409, detail=f"指标 '{req.metric_code}' 已存在")
        store.delete_metric(metric_code)
        metric.metric_code = req.metric_code
        metric_code = req.metric_code

    if req.object_code is not None and req.object_code != metric.object_code:
        if not store.get_object(req.object_code):
            raise HTTPException(status_code=400, detail=f"对象 '{req.object_code}' 不存在")
        # Sync metric_code prefix with new object_code
        old_code = metric.metric_code
        old_prefix = metric.object_code
        suffix = old_code[len(old_prefix):] if old_code.startswith(old_prefix + ".") else old_code.split(".", 1)[-1] if "." in old_code else old_code
        new_code = f"{req.object_code}.{suffix}"
        if new_code != old_code and store.get_metric(new_code):
            raise HTTPException(status_code=409, detail=f"指标 '{new_code}' 已存在")
        store.delete_metric(old_code)
        metric.object_code = req.object_code
        metric.metric_code = new_code

    if req.name is not None:
        metric.name = req.name
    if req.definition is not None:
        metric.definition = req.definition
    if req.metric_type is not None:
        metric.metric_type = req.metric_type
    if req.semantic_type is not None:
        metric.semantic_type = req.semantic_type
    if req.unit is not None:
        metric.unit = req.unit
    if req.importance is not None:
        metric.importance = req.importance
    if req.source_field is not None:
        metric.source_field = req.source_field
        # Auto-calculate quality_score from discovery field metadata
        metric.quality_score = _calc_quality_from_discovery(req.source_field, metric.source_object)
    if req.source_object is not None:
        metric.source_object = req.source_object
    if req.source_adapter is not None:
        metric.source_adapter_port = req.source_adapter
    if req.value_domain is not None:
        metric.value_domain = req.value_domain
    if req.required is not None:
        metric.required = req.required

    store.save_metric(metric)
    return {"status": "ok", "metric_code": metric.metric_code}


@router.delete("/metrics/{metric_code:path}")
def delete_metric(metric_code: str):
    """删除业务指标。"""
    reg = get_registry()
    store = reg._store
    metric = store.get_metric(metric_code)
    if not metric:
        raise HTTPException(status_code=404, detail=f"指标 '{metric_code}' 不存在")
    store.delete_metric(metric_code)
    return {"status": "ok", "metric_code": metric_code}


@router.get("/health")
def health_check():
    reg = get_registry()
    return {"status": "ok", "objects_count": len(reg.list_objects()), "store_type": "in_memory"}


class CreateMetricRequest(BaseModel):
    object_code: str
    name: str
    metric_type: str = "Atomic"
    semantic_type: str = "Amount"
    definition: str | None = None
    unit: str | None = None
    importance: str = "optional"
    value_domain: str | None = None
    required: bool = False
    source_table: str | None = None
    source_field: str | None = None


class UpdateMetricRequest(BaseModel):
    metric_code: str | None = None
    object_code: str | None = None
    name: str | None = None
    definition: str | None = None
    metric_type: str | None = None
    semantic_type: str | None = None
    unit: str | None = None
    importance: str | None = None
    source_field: str | None = None
    source_object: str | None = None
    source_adapter: str | None = None
    source_table: str | None = None
    value_domain: str | None = None
    required: bool | None = None


class CreateDomainRequest(BaseModel):
    domain_code: str
    name: str


class UpdateDomainRequest(BaseModel):
    name: str


class DomainInfo(BaseModel):
    domain_code: str
    name: str


@router.get("/domains", response_model=list[DomainInfo])
def list_domains():
    """列出所有业务域（供筛选下拉等使用）。"""
    reg = get_registry()
    store = reg._store
    return [DomainInfo(domain_code=d.domain_code, name=d.name) for d in store.list_domains()]


@router.post("/domains")
def create_domain(req: CreateDomainRequest):
    """创建新的业务域。"""
    from src.semantic_layer.models import BusinessDomain
    reg = get_registry()
    store = reg._store
    existing = store.get_domain(req.domain_code)
    if existing:
        raise HTTPException(status_code=409, detail=f"域 '{req.domain_code}' 已存在")
    domain = BusinessDomain(domain_code=req.domain_code, name=req.name)
    store.save_domain(domain)
    return {"status": "ok", "domain_code": req.domain_code, "name": req.name}


@router.put("/domains/{domain_code}")
def update_domain(domain_code: str, req: UpdateDomainRequest):
    """更新业务域名称。"""
    reg = get_registry()
    store = reg._store
    domain = store.get_domain(domain_code)
    if not domain:
        raise HTTPException(status_code=404, detail=f"域 '{domain_code}' 不存在")
    domain.name = req.name
    store.save_domain(domain)
    return {"status": "ok", "domain_code": domain_code, "name": req.name}


@router.delete("/domains/{domain_code}")
def delete_domain(domain_code: str):
    """删除业务域。不能被 skill 引用的域不允许删除。"""
    reg = get_registry()
    store = reg._store
    domain = store.get_domain(domain_code)
    if not domain:
        raise HTTPException(status_code=404, detail=f"域 '{domain_code}' 不存在")

    # 检查是否有 skill 引用了该域下的指标
    domain_objects = [o for o in store.list_objects() if o.domain_code == domain_code]
    blocked_metrics: list[str] = []
    for obj in domain_objects:
        for metric in store.list_metrics(object_code=obj.object_code):
            if metric.usage_count > 0:
                blocked_metrics.append(metric.metric_code)

    if blocked_metrics:
        raise HTTPException(
            status_code=409,
            detail=f"无法删除：以下指标被 {len(blocked_metrics)} 个 Skill 引用，请先解除引用后再删除域："
                   f"{', '.join(blocked_metrics[:5])}"
                   f"{'…等' if len(blocked_metrics) > 5 else ''}",
        )

    store.delete_domain(domain_code)
    return {"status": "ok", "domain_code": domain_code}


# ── 技能引用计数（静态计算）────────────────────────────────────
# 背景：usage_count 原为运行时计数器，仅在技能执行时累加；未执行时恒为 0，
# 看板「技能引用」无法反映技能对指标的静态引用关系。
# 改为静态扫描 skill_manifest：按 metric_code（生产环境）或 source_field（内存种子）
# 双策略匹配，统计每个指标被多少 Skill 引用。
_skill_refs_cache: dict[str, int] | None = None
_skill_refs_cache_ts: float = 0.0
_SKILL_REFS_TTL = 30.0  # 缓存 30s，避免 metrics 页面 N+1 调用重复扫描


def _compute_skill_metric_refs() -> dict[str, int]:
    """静态计算每个指标被多少 Skill 引用。

    needed_objects 声明的 {object_code}.{metric} 直接匹配语义层 metric_code。
    编码统一后（skill_manifest 与语义层均为 zydyxx.* 物理编码），无需 source_field 兼容。
    返回 {metric_code: 引用该指标的 Skill 数}。
    """
    import yaml
    from pathlib import Path
    from src.config.production import SKILLS_DIR

    reg = get_registry()
    store = reg._store
    all_metrics = store.list_metrics()

    if not all_metrics:
        return {}

    all_metric_codes = {m.metric_code for m in all_metrics}

    skills_root = Path(SKILLS_DIR)
    refs: dict[str, int] = {}

    for manifest_path in skills_root.glob("*/skill_manifest.yaml"):
        try:
            with open(manifest_path, "r", encoding="utf-8") as f:
                manifest = yaml.safe_load(f) or {}
        except Exception:
            logger.warning("读取 skill_manifest 失败: %s", manifest_path, exc_info=True)
            continue

        # 命中的指标（同一 Skill 对同一指标只计一次）
        hit_metrics: set[str] = set()

        # needed_objects 的 {object_code}.{metric} 直接匹配语义层 metric_code
        for obj_decl in manifest.get("needed_objects", []):
            obj_code = obj_decl.get("object_code", "")
            for mc in obj_decl.get("metrics", []):
                full_code = f"{obj_code}.{mc}"
                if full_code in all_metric_codes:
                    hit_metrics.add(full_code)

        for mc in hit_metrics:
            refs[mc] = refs.get(mc, 0) + 1

    return refs


def _get_skill_metric_refs() -> dict[str, int]:
    """带 TTL 缓存地获取技能引用计数。"""
    global _skill_refs_cache, _skill_refs_cache_ts
    now = time.time()
    if _skill_refs_cache is not None and (now - _skill_refs_cache_ts) < _SKILL_REFS_TTL:
        return _skill_refs_cache
    _skill_refs_cache = _compute_skill_metric_refs()
    _skill_refs_cache_ts = now
    return _skill_refs_cache


def _compute_skill_locks() -> dict[str, list[dict]]:
    """扫描 skill_manifest 的 locked_versions，返回每个对象的锁定情况。

    返回 {object_code: [{skill_id, locked_version}, ...]}。
    locked_version=None 表示跟随最新已发布（follow latest published）。
    """
    import yaml
    from pathlib import Path
    from src.config.production import SKILLS_DIR

    locks: dict[str, list[dict]] = {}
    for manifest_path in Path(SKILLS_DIR).glob("*/skill_manifest.yaml"):
        try:
            with open(manifest_path, encoding="utf-8") as f:
                manifest = yaml.safe_load(f) or {}
        except Exception:
            logger.warning("读取 skill_manifest 失败: %s", manifest_path, exc_info=True)
            continue
        skill_id = manifest.get("skill_id", manifest_path.parent.name)
        for obj_code, ver in (manifest.get("locked_versions") or {}).items():
            locks.setdefault(obj_code, []).append(
                {"skill_id": skill_id, "locked_version": ver})
    return locks


@router.get("/summary", response_model=SemanticSummary)
def get_semantic_summary():
    reg = get_registry()
    store = reg._store

    domains = store.list_domains()
    all_objects = store.list_objects()
    all_metrics = store.list_metrics()
    skill_refs_map = _get_skill_metric_refs()

    domains_count = len(domains)
    objects_count = len(all_objects)
    metrics_count = len(all_metrics)

    def _is_mapped(m) -> bool:
        """与前端 determineMappingStatus 对齐"""
        if not m.source_field:
            return False
        return not (m.semantic_type == 'Enum' and not m.value_domain)

    mapped_count = sum(1 for m in all_metrics if _is_mapped(m))
    value_missing_count = sum(1 for m in all_metrics if m.source_field and m.semantic_type == 'Enum' and not m.value_domain)
    unmapped_count = sum(1 for m in all_metrics if not m.source_field)
    mapping_rate = (mapped_count / metrics_count * 100.0) if metrics_count > 0 else 0.0

    skill_references = sum(skill_refs_map.values())

    domain_progress = []
    for domain in domains:
        domain_objects = [o for o in all_objects if o.domain_code == domain.domain_code]
        domain_object_codes = [o.object_code for o in domain_objects]

        domain_metrics = [m for m in all_metrics if m.object_code in domain_object_codes]
        total = len(domain_metrics)
        mapped = sum(1 for m in domain_metrics if _is_mapped(m))
        pct = (mapped / total * 100.0) if total > 0 else 0.0
        refs = sum(skill_refs_map.get(m.metric_code, 0) for m in domain_metrics)
        domain_progress.append(DomainProgress(
            domain_code=domain.domain_code,
            name=domain.name,
            total_metrics=total,
            mapped_metrics=mapped,
            percentage=round(pct, 1),
            skill_refs=refs,
        ))

    # 发现扫描汇总：取最近一次成功扫描的表/字段统计，无扫描记录时降级为 0
    discovery_tables = 0
    discovery_fields = 0
    discovery_unmapped = 0
    try:
        latest_scan = _get_discovery_store().get_latest_result()
        if latest_scan:
            discovery_tables = latest_scan.get("total_tables", 0) or 0
            discovery_fields = latest_scan.get("total_fields", 0) or 0
            discovery_unmapped = latest_scan.get("unmapped_fields", 0) or 0
    except Exception:
        logger.warning("get_semantic_summary: 加载发现扫描结果失败，降级为 0", exc_info=True)

    return SemanticSummary(
        domains_count=domains_count,
        objects_count=objects_count,
        metrics_count=metrics_count,
        mapped_count=mapped_count,
        unmapped_count=unmapped_count,
        value_missing_count=value_missing_count,
        mapping_rate=round(mapping_rate, 1),
        skill_references=skill_references,
        domain_progress=domain_progress,
        discovery_tables=discovery_tables,
        discovery_fields=discovery_fields,
        discovery_unmapped=discovery_unmapped,
    )


def _calc_recency(last_updated: str | None) -> int:
    """计算活跃度得分（15/13/10/6/2/0）。"""
    if not last_updated:
        return 0
    try:
        from datetime import datetime, timezone
        dt = datetime.fromisoformat(last_updated.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        days = (datetime.now(timezone.utc) - dt).days
        if days <= 7: return 15
        if days <= 30: return 13
        if days <= 90: return 10
        if days <= 180: return 6
        if days <= 365: return 2
        return 0
    except Exception:
        return 0


def _calc_usage(count: int) -> int:
    """计算使用度得分（15/12/8/4/0）。"""
    if count >= 20: return 15
    if count >= 10: return 12
    if count >= 5: return 8
    if count >= 1: return 4
    return 0


def calc_field_value_score(field: dict, usage_count: int = 0) -> ValueScoreDetail:
    """五维数据价值评分（0-100）。非空率×50 + 描述0/5/10 + 示例0/10 + 活跃度0-15 + 使用度0-15。"""
    nnr = float(field.get("non_null_rate", 0) or 0)
    # Normalize: if >1, treat as percentage (0-100)
    if nnr > 1:
        nnr = nnr / 100.0

    desc = (field.get("description") or "").strip()
    sv = field.get("sample_value")
    sv_list = field.get("sample_values")
    has_sample = bool(
        (sv and str(sv).lower() != "null" and str(sv).strip() != "")
        or (sv_list and len(sv_list) > 0 and any(s and str(s).lower() != "null" for s in sv_list))
    )
    lu = field.get("last_updated")

    non_null_score = round(nnr * 50, 1)
    desc_score = 10 if len(desc) >= 4 else (5 if desc else 0)
    sample_score = 10 if has_sample else 0
    recency_score = _calc_recency(lu)
    usage_score = _calc_usage(usage_count)

    total = non_null_score + desc_score + sample_score + recency_score + usage_score
    grade = "A" if total >= 85 else "B" if total >= 70 else "C" if total >= 50 else "D" if total >= 30 else "E"

    return ValueScoreDetail(
        total=total, grade=grade,
        non_null_score=non_null_score, non_null_rate=round(nnr, 4),
        desc_score=desc_score, has_desc=bool(desc),
        sample_score=sample_score, has_sample=has_sample,
        recency_score=recency_score, last_updated=lu,
        usage_score=usage_score, usage_count=usage_count,
    )


def calc_field_quality(field: dict, usage_count: int = 0) -> float:
    """旧接口兼容 — 返回 total 值。"""
    return calc_field_value_score(field, usage_count).total


def _calc_quality_from_discovery(source_field: str, source_object: str | None = None) -> float:
    """根据发现中心字段元数据自动计算质量分。"""
    try:
        from src.data_platform.storage.postgresql.discovery_store import DiscoveryStore
        store = DiscoveryStore()
        latest = store.get_latest_result()
        if not latest:
            return 0.0
        field_name = source_field
        table_name = source_object or ""
        if "." in source_field:
            parts = source_field.split(".", 1)
            table_name = parts[0]
            field_name = parts[1]
        for f in latest.get("fields", []):
            fn = f.get("field_name", "")
            tn = f.get("table_name", "")
            if fn == field_name and (not table_name or tn == table_name):
                return calc_field_value_score(f).total
    except Exception as e:
        import logging; logging.getLogger(__name__).warning("_calc_quality_from_discovery failed for '%s': %s", source_field, e)
    return 0.0


def _refresh_quality_scores_from_scan(result_fields: list[dict]) -> int:
    """扫描完成后，根据最新字段元数据批量刷新已映射指标的质量分。

    替代原先 GET /metrics/{code} 里“读时算+写”的反模式：质量分在扫描
    结束后一次性同步入库，详情接口只需返回已持久化的值。
    Returns: 实际更新的指标条数。
    """
    reg = get_registry()
    store = reg._store
    try:
        metrics = store.list_metrics()
    except Exception:
        logger.warning("_refresh_quality_scores_from_scan: list_metrics 失败", exc_info=True)
        return 0

    # 构建字段查找索引：优先 table.field 全路径，回退纯字段名
    by_full: dict[str, dict] = {}
    by_name: dict[str, dict] = {}
    for f in result_fields:
        fn = (f.get("field_name") or "")
        tn = (f.get("table_name") or "")
        if fn:
            by_full[f"{tn}.{fn}".lower()] = f
            by_name[fn.lower()] = f

    updated = 0
    for m in metrics:
        if not m.source_field:
            continue
        sf = m.source_field.lower().strip()
        field_meta = None
        if "." in sf:
            field_meta = by_full.get(sf)
            if field_meta is None:
                field_meta = by_name.get(sf.split(".", 1)[1])
        else:
            field_meta = by_name.get(sf)
            if field_meta is None and m.source_object:
                field_meta = by_full.get(f"{m.source_object.lower()}.{sf}")
        qs = calc_field_value_score(field_meta).total if field_meta else 0.0
        if qs != m.quality_score:
            m.quality_score = qs
            store.save_metric(m)
            updated += 1
    if updated:
        logger.info("_refresh_quality_scores_from_scan: 刷新 %d 个指标质量分", updated)
    return updated


# ── Discovery 字段语义匹配：Milvus 单例（避免每请求重连） ──
_discovery_milvus: dict | None = None  # {"collection": col, "provider": provider}


def _get_discovery_milvus() -> dict | None:
    """懒加载 discovery_fields Milvus 集合（应用级单例，连接仅建立一次）。

    不可用（未安装 pymilvus / 未配置 / 集合不存在）时返回 None，调用方降级。
    """
    global _discovery_milvus
    if _discovery_milvus is not None:
        return _discovery_milvus
    try:
        from pymilvus import Collection, connections, utility
        from src.config.production import MILVUS_HOST, MILVUS_PORT
        from src.knowledge_extension.rule_explanation.policy_retrieval.embedding_provider import get_embedding_provider

        # 使用独立 alias，避免与其他模块的 default 连接冲突
        connections.connect(alias="discovery", host=MILVUS_HOST, port=str(MILVUS_PORT))
        if not utility.has_collection("discovery_fields"):
            logger.info("field_match: Milvus 无 discovery_fields 集合，降级为模糊匹配")
            return None
        col = Collection("discovery_fields")
        col.load()
        provider = get_embedding_provider("sentence_transformer")
        _discovery_milvus = {"collection": col, "provider": provider}
        return _discovery_milvus
    except Exception:
        logger.warning("field_match: Milvus 初始化失败，降级为模糊匹配", exc_info=True)
        return None


def _reset_discovery_milvus() -> None:
    """重置 Milvus 单例，强制下次调用重连。"""
    global _discovery_milvus
    _discovery_milvus = None


class FieldMatchRequest(BaseModel):
    query: str
    definition: str = ""


class FieldMatchItem(BaseModel):
    field_name: str
    table_name: str
    description: str = ""
    score: float = 0.0


class FieldMatchResponse(BaseModel):
    matches: list[FieldMatchItem]


@router.post("/field-match", response_model=FieldMatchResponse)
def field_match(req: FieldMatchRequest):
    """AI 语义匹配：根据指标中文名+定义，从发现中心字段中匹配最相似的字段。"""
    query_text = f"{req.query} {req.definition}".strip()
    matches: list[FieldMatchItem] = []

    # 优先使用 Milvus 向量检索（复用应用级单例连接）
    milvus = _get_discovery_milvus()
    if milvus is not None:
        try:
            col = milvus["collection"]
            provider = milvus["provider"]
            query_vec = provider.encode([query_text])[0].tolist()
            results = col.search(
                data=[query_vec], anns_field="embedding",
                param={"metric_type": "COSINE", "params": {"ef": 64}},
                limit=5, output_fields=["field_name", "table_name", "description"],
            )
            for hits in results:
                for h in hits:
                    matches.append(FieldMatchItem(
                        field_name=h.entity.get("field_name", ""),
                        table_name=h.entity.get("table_name", ""),
                        description=h.entity.get("description", ""),
                        score=round(float(h.distance), 4),
                    ))
            if matches:
                return FieldMatchResponse(matches=matches)
        except Exception:
            logger.warning("field_match: Milvus 查询失败，降级为模糊匹配", exc_info=True)
            # 连接可能已断，重置单例以便下次重试
            _reset_discovery_milvus()

    # Fallback: fuzzy text match from discovery results
    from src.data_platform.storage.postgresql.discovery_store import DiscoveryStore
    store = DiscoveryStore()
    latest = store.get_latest_result()
    if latest:
        fields = latest.get("fields", [])
        q = query_text.lower()
        scored = []
        for f in fields:
            fn = f.get("field_name", "").lower()
            desc = (f.get("description") or "").lower()
            score = 0.0
            if q in fn: score += 0.5
            if q in desc: score += 0.3
            # Partial word match
            for w in q.split():
                if w in fn: score += 0.1
                if w in desc: score += 0.05
            if score > 0:
                scored.append((score, f))
        scored.sort(key=lambda x: -x[0])
        for s, f in scored[:5]:
            matches.append(FieldMatchItem(
                field_name=f.get("field_name", ""),
                table_name=f.get("table_name", ""),
                description=f.get("description", ""),
                score=round(s, 4),
            ))

    return FieldMatchResponse(matches=matches)


class ValueDomainInfo(BaseModel):
    domain_code: str
    name: str
    description: str | None = None
    mapping_count: int = 0
    standard_values: list[str] = []


class StandardValuesRequest(BaseModel):
    standard_values: list[str]


class ValueDomainMappingsResponse(BaseModel):
    domain_code: str
    mappings: list[ValueMappingResponse]


@router.get("/value-domains", response_model=list[ValueDomainInfo])
def list_value_domains():
    """列出所有值域及其映射数量。"""
    reg = get_registry()
    store = reg._store
    # PostgreSQL：批量 JOIN 查询，避免 N+1
    if hasattr(store, "list_value_domains_with_counts"):
        return [
            ValueDomainInfo(
                domain_code=vd.domain_code, name=vd.name, description=vd.description,
                mapping_count=cnt, standard_values=vd.standard_values or [],
            )
            for vd, cnt in store.list_value_domains_with_counts()
        ]
    # 内存版回退（值域数量少，逐条可接受）
    result = []
    if hasattr(store, "_value_domains"):
        for vd_code in list(store._value_domains.keys()):
            vd = store.get_value_domain(vd_code)
            if vd:
                mappings = store.get_value_mappings(vd_code)
                result.append(ValueDomainInfo(domain_code=vd.domain_code, name=vd.name, description=vd.description, mapping_count=len(mappings), standard_values=vd.standard_values or []))
    return result


class CreateValueDomainRequest(BaseModel):
    domain_code: str
    name: str
    description: str | None = None


@router.post("/value-domains")
def create_value_domain(req: CreateValueDomainRequest):
    """创建新的值域。"""
    from src.semantic_layer.models import ValueDomain
    reg = get_registry()
    store = reg._store
    if store.get_value_domain(req.domain_code):
        raise HTTPException(status_code=409, detail=f"值域 '{req.domain_code}' 已存在")
    vd = ValueDomain(domain_code=req.domain_code, name=req.name, description=req.description)
    store.save_value_domain(vd)
    return {"status": "ok", "domain_code": req.domain_code, "name": req.name}


@router.delete("/value-domains/{domain_code}")
def delete_value_domain(domain_code: str):
    """删除值域及其所有映射。"""
    reg = get_registry()
    store = reg._store
    if not store.get_value_domain(domain_code):
        raise HTTPException(status_code=404, detail=f"值域 '{domain_code}' 不存在")
    store.delete_value_domain(domain_code)
    return {"status": "ok", "domain_code": domain_code}


@router.get("/value-domains/{domain_code}/standard-values")
def get_standard_values(domain_code: str):
    """获取值域的标准值列表。"""
    reg = get_registry()
    store = reg._store
    vd = store.get_value_domain(domain_code)
    if not vd:
        raise HTTPException(status_code=404, detail=f"值域 '{domain_code}' 不存在")
    return {"domain_code": domain_code, "standard_values": vd.standard_values or []}


@router.put("/value-domains/{domain_code}/standard-values")
def update_standard_values(domain_code: str, req: StandardValuesRequest):
    """更新值域的标准值列表（全量替换）。"""
    reg = get_registry()
    store = reg._store
    vd = store.get_value_domain(domain_code)
    if not vd:
        raise HTTPException(status_code=404, detail=f"值域 '{domain_code}' 不存在")
    vd.standard_values = req.standard_values
    store.save_value_domain(vd)
    return {"status": "ok", "domain_code": domain_code, "count": len(req.standard_values)}


@router.get("/value-domains/{domain_code}/mappings", response_model=ValueDomainMappingsResponse)
def get_value_domain_mappings(domain_code: str):
    """获取指定值域的所有映射。"""
    reg = get_registry()
    store = reg._store
    vd = store.get_value_domain(domain_code)
    if not vd:
        raise HTTPException(status_code=404, detail=f"值域 '{domain_code}' 不存在")
    mappings = store.get_value_mappings(domain_code)
    return ValueDomainMappingsResponse(
        domain_code=domain_code,
        mappings=[ValueMappingResponse(status="ok", domain_code=m.domain_code, source_value=m.source_value, standard_value=m.standard_value) for m in mappings],
    )


@router.post("/value-domain/mapping", response_model=ValueMappingResponse)
def save_value_mapping(req: ValueMappingRequest):
    """Save a source_value → standard_value mapping."""
    reg = get_registry()
    store = reg._store
    from src.semantic_layer.models import ValueDomainMapping
    mapping = ValueDomainMapping(
        id=None,
        domain_code=req.domain_code,
        source_value=req.source_value,
        standard_value=req.standard_value,
        description="",
    )
    store.save_value_mapping(mapping)
    return ValueMappingResponse(
        status="ok",
        domain_code=req.domain_code,
        source_value=req.source_value,
        standard_value=req.standard_value,
    )


@router.delete("/value-domains/{domain_code}/mappings/{source_value:path}")
def delete_value_mapping(domain_code: str, source_value: str):
    """删除单个值域映射。"""
    reg = get_registry()
    store = reg._store
    if not store.get_value_domain(domain_code):
        raise HTTPException(status_code=404, detail=f"值域 '{domain_code}' 不存在")
    store.delete_value_mapping(domain_code, source_value)
    return {"status": "ok", "domain_code": domain_code, "source_value": source_value}


# ── 指标引用追踪 ──────────────────────────────────────────────────


class TrackUsageResponse(BaseModel):
    metric_code: str
    usage_count: int


class SkillMetricItem(BaseModel):
    metric_code: str
    name: str
    object_code: str
    usage_count: int


class UsageStatsResponse(BaseModel):
    total_metrics: int
    total_usage: int
    top_metrics: list[SkillMetricItem]


@router.post("/metrics/{metric_code:path}/track-usage", response_model=TrackUsageResponse)
def track_metric_usage(metric_code: str):
    """技能引用指标时调用，usage_count +1。"""
    reg = get_registry()
    store = reg._store
    metric = store.get_metric(metric_code)
    if not metric:
        raise HTTPException(status_code=404, detail=f"指标 '{metric_code}' 不存在")
    metric.usage_count = (metric.usage_count or 0) + 1
    store.save_metric(metric)
    return TrackUsageResponse(metric_code=metric_code, usage_count=metric.usage_count)


@router.get("/metrics/usage-stats", response_model=UsageStatsResponse)
def get_usage_stats():
    """获取全局指标引用统计（引用排行 TOP 20）。"""
    reg = get_registry()
    store = reg._store
    all_metrics = store.list_metrics()
    used = [m for m in all_metrics if m.usage_count > 0]
    used.sort(key=lambda m: m.usage_count, reverse=True)
    return UsageStatsResponse(
        total_metrics=len(all_metrics),
        total_usage=sum(m.usage_count for m in all_metrics),
        top_metrics=[SkillMetricItem(
            metric_code=m.metric_code, name=m.name, object_code=m.object_code,
            usage_count=m.usage_count,
        ) for m in used[:20]],
    )


@router.get("/skills/{skill_id}/metrics", response_model=list[SkillMetricItem])
def get_skill_metrics(skill_id: str):
    """返回指定技能引用的所有指标（从 manifest 解析）。"""
    import yaml
    from pathlib import Path
    from src.config.production import SKILLS_DIR
    manifest_path = Path(SKILLS_DIR) / skill_id / "skill_manifest.yaml"
    if not manifest_path.exists():
        raise HTTPException(status_code=404, detail=f"技能 '{skill_id}' 不存在")
    with open(manifest_path, "r", encoding="utf-8") as f:
        manifest = yaml.safe_load(f)
    reg = get_registry()
    store = reg._store
    result: list[SkillMetricItem] = []
    for obj_decl in manifest.get("needed_objects", []):
        obj_code = obj_decl.get("object_code", "")
        for metric_code in obj_decl.get("metrics", []):
            full_code = f"{obj_code}.{metric_code}"
            metric = store.get_metric(full_code)
            if metric:
                result.append(SkillMetricItem(
                    metric_code=metric.metric_code, name=metric.name,
                    object_code=metric.object_code, usage_count=metric.usage_count,
                ))
    return result


def _resolve_skill_metric_codes(skill_id: str) -> list[str]:
    """从 skill_manifest.yaml 的 needed_objects 解析出完整指标编码列表。"""
    import yaml
    from pathlib import Path
    from src.config.production import SKILLS_DIR
    manifest_path = Path(SKILLS_DIR) / skill_id / "skill_manifest.yaml"
    if not manifest_path.exists():
        raise HTTPException(status_code=404, detail=f"技能 '{skill_id}' 不存在")
    with open(manifest_path, "r", encoding="utf-8") as f:
        manifest = yaml.safe_load(f) or {}
    codes: list[str] = []
    for obj_decl in manifest.get("needed_objects", []):
        obj_code = obj_decl.get("object_code", "")
        for m in obj_decl.get("metrics", []):
            codes.append(f"{obj_code}.{m}")
    return codes


@router.get("/skills/{skill_id}/query-plan")
def get_skill_query_plan(skill_id: str):
    """返回技能的取数查询计划（纯元数据，不执行 SQL）。

    展示：技能消费的指标 → 解析为物理表/列 → 按表分组的批量取数计划 → 未映射项。
    供 skills 页「查询计划」Tab 可视化取数透视。
    """
    from src.runtime.discovery.semantic_source import get_semantic_data_source

    metric_codes = _resolve_skill_metric_codes(skill_id)
    if not metric_codes:
        raise HTTPException(status_code=404, detail=f"技能 '{skill_id}' 未声明 needed_objects")
    source = get_semantic_data_source()
    plan = source.build_query_plan(metric_codes)
    return plan


class SkillQueryExecuteRequest(BaseModel):
    djh: str | int
    patient_id: str | None = None
    encounter_id: str | None = None


@router.post("/skills/{skill_id}/query-execute")
def execute_skill_query(skill_id: str, req: SkillQueryExecuteRequest):
    """试运行：按给定 djh 真实取数，返回每个指标的实际值。

    复用 discovery 的 SQL Server 连接通道。
    """
    from src.runtime.discovery.semantic_source import get_semantic_data_source

    metric_codes = _resolve_skill_metric_codes(skill_id)
    source = get_semantic_data_source()
    context = {"djh": req.djh}
    if req.patient_id:
        context["patient_id"] = req.patient_id
    if req.encounter_id:
        context["encounter_id"] = req.encounter_id
    values = source.query(metric_codes, context=context)
    # 附带 source_field 便于前端展示「指标 → 物理字段 → 值」
    reg = get_registry()
    store = reg._store
    items = []
    for code in metric_codes:
        m = store.get_metric(code)
        items.append({
            "metric_code": code,
            "name": m.name if m else code,
            "source_field": m.source_field if m else None,
            "value": values.get(code),
        })
    return {"skill_id": skill_id, "djh": req.djh, "items": items}


# SettlementContext 字段 → 语义指标编码（一致性校验用）
# [来源: settlement_data_provider.SettlementContext 与语义层指标的对应]
_SETTLEMENT_CONTEXT_TO_METRIC = {
    "deductible": "zydyxx.bcqfje",
    "medical_insurance_inner_amount": "zydyxx.bcybnje",
    "basic_pooling_payment": "zyfdxx.bdtczfje",
    "basic_pooling_self_pay": "zyfdxx.bdtczf",
    "large_amount_payment": "zyfdxx.bddezfje",
    "large_amount_self_pay": "zyfdxx.bddezf",
    "personal_total_pay": "zyfdxx.bdgryf",
    "person_type": "zyjyxx.rylb",
    "insurance_type": "djxx.fund_type",
    "service_type": "djxx.yllb",
}


@router.get("/skills/{skill_id}/consistency-check")
def consistency_check_skill(skill_id: str, djh: str):
    """一致性校验：同一 djh 下，语义层路径 vs 现有 business_sql 路径的取数对比。

    回答“语义层建的映射对不对”：对每个指标对比两条路径的值，标出差异。
    仅 settlement_explain_skill 有并行 business_sql 路径；其他技能返回未支持。
    """
    from src.runtime.discovery.semantic_source import get_semantic_data_source

    if skill_id != "settlement_explain_skill":
        return {"skill_id": skill_id, "supported": False,
                "message": "该技能无并行 business_sql 路径，无法对比"}

    metric_codes = _resolve_skill_metric_codes(skill_id)

    # ① 语义层 flat 路径
    source = get_semantic_data_source()
    semantic_values = source.query(metric_codes, context={"djh": djh})
    # ①' 语义层 joined 路径（复用 business_sql JOIN）
    joined_values = source.query(metric_codes, context={"djh": djh}, join_mode="joined")

    # ② business_sql 路径（现有生产路径）
    business_values: dict[str, Any] = {}
    business_error: str | None = None
    try:
        from src.config.production import DATA_SOURCE_MODE
        if DATA_SOURCE_MODE != "real_db":
            business_error = f"DATA_SOURCE_MODE={DATA_SOURCE_MODE}，未启用真实 DB"
        else:
            from src.runtime.policy_qa.settlement_data_provider import RealDbSettlementDataProvider
            provider = RealDbSettlementDataProvider()
            raw = provider.client.get_case_context_raw(settlement_id=djh)
            raw_data = raw.raw_data or {}
            # 映射 SettlementContext 字段 → 指标编码
            for ctx_field, metric_code in _SETTLEMENT_CONTEXT_TO_METRIC.items():
                raw_key = {
                    "deductible": "bcqfje", "medical_insurance_inner_amount": "bcybnje",
                    "basic_pooling_payment": "bdtczfje", "basic_pooling_self_pay": "bdtczf",
                    "large_amount_payment": "bddegwyzfje", "large_amount_self_pay": "bddegwyzf",
                    "personal_total_pay": "bdgryf", "person_type": "PER_TYPE",
                    "insurance_type": "fund_type", "service_type": "yllb",
                }.get(ctx_field, ctx_field)
                if raw_key in raw_data:
                    business_values[metric_code] = raw_data.get(raw_key)
    except Exception as e:
        business_error = str(e)

    # ③ 逐指标对比
    reg = get_registry()
    store = reg._store
    items = []
    for code in metric_codes:
        m = store.get_metric(code)
        sem_v = semantic_values.get(code)
        join_v = joined_values.get(code)
        biz_v = business_values.get(code)
        # 只对两条路径都返回的数值型指标判定 match
        compared = code in business_values
        match = (compared and _values_equal(sem_v, biz_v))
        joined_match = (compared and _values_equal(join_v, biz_v))
        items.append({
            "metric_code": code,
            "name": m.name if m else code,
            "semantic_value": sem_v,
            "semantic_joined_value": join_v,
            "business_sql_value": biz_v,
            "compared": compared,
            "match": match,
            "joined_match": joined_match,
        })
    matched = sum(1 for it in items if it["match"])
    joined_matched = sum(1 for it in items if it["joined_match"])
    compared = sum(1 for it in items if it["compared"])
    return {
        "skill_id": skill_id, "djh": djh, "supported": True,
        "business_sql_error": business_error,
        "summary": {
            "compared": compared,
            "flat_matched": matched, "flat_mismatched": compared - matched,
            "joined_matched": joined_matched, "joined_mismatched": compared - joined_matched,
        },
        "items": items,
    }


def _values_equal(a: Any, b: Any) -> bool:
    """数值型容差比较；非数值要求相等。"""
    try:
        if isinstance(a, (int, float)) and isinstance(b, (int, float)):
            return abs(float(a) - float(b)) < 0.01
    except Exception:
        pass
    return str(a) == str(b)



# ── Discovery scan endpoints ────────────────────────────────────────────────────────


def _collect_source_fields() -> set[str]:
    """Return all source_field values from registry metrics (lowercased)."""
    reg = get_registry()
    store = reg._store
    return {m.source_field.lower() for m in store.list_metrics() if m.source_field}


class DiscoveryScanRequest(BaseModel):
    task_id: str | None = None
    source_config: dict | None = None
    scope: str = "全部已接入表"
    sample_limit: int = 10000  # 样本值行数上限（页面可配置）


# ── Discovery store（PostgreSQL 持久化，重启不丢） ──────────────────────────

_discovery_store: "DiscoveryStore | None" = None


def _get_discovery_store() -> "DiscoveryStore":
    """延迟初始化 DiscoveryStore 单例。"""
    global _discovery_store
    if _discovery_store is None:
        from src.data_platform.storage.postgresql.discovery_store import DiscoveryStore
        _discovery_store = DiscoveryStore()
    return _discovery_store


def _run_discovery_sync(source_config: dict, store=None) -> dict:
    """Lazy wrapper to avoid circular import at module load time."""
    from src.runtime.discovery.service import run_discovery
    return run_discovery(source_config, store)


@router.post("/discovery/incremental-update", response_model=DiscoveryScanResponse)
def incremental_discovery_update():
    """增量更新：刷新字段 last_updated、检测新增字段、重算价值分。"""
    task_id = str(uuid.uuid4())
    store = _get_discovery_store()
    store.create_task(task_id, {"mode": "incremental"}, sample_limit=0)

    latest = store.get_latest_result()
    if not latest:
        raise HTTPException(status_code=400, detail="尚未执行过全量扫描，请先点击「重新扫描」")

    from datetime import datetime, timezone
    now = datetime.now(timezone.utc).isoformat()

    fields = latest.get("fields", [])
    original_len = len(fields)
    updated_count = 0

    # 更新现有字段的 last_updated + 重算价值分
    for f in fields:
        old_lu = f.get("last_updated")
        # 保留原始 last_updated（来自真实扫描），否则设为当前时间
        if not old_lu:
            f["last_updated"] = now
            updated_count += 1

    # 模拟新增字段检测（真实场景会从源数据库扫描新列）
    existing_names = {f["field_name"] for f in fields}
    for tbl, cols in _MOCK_FIELD_METADATA.items():
        for col_name, meta in cols.items():
            if col_name not in existing_names:
                fields.append({
                    "field_name": col_name,
                    "table_name": tbl,
                    "description": meta.description,
                    "data_type": meta.data_type,
                    "non_null_rate": meta.non_null_rate,
                    "non_null_row_count": meta.non_null_row_count,
                    "distinct_count": meta.distinct_count,
                    "sample_value": meta.sample_value,
                    "sample_values": meta.sample_values,
                    "sample_stats": None,
                    "is_dictionary": meta.is_dictionary,
                    "last_updated": now,
                    "suggested_object": None,
                    "mapped": False,
                    "table_schema": None,
                    "is_nullable": meta.is_nullable,
                    "is_primary_key": meta.is_primary_key,
                    "remark": None,
                })
                existing_names.add(col_name)

    new_count = len(fields) - original_len

    result = {
        "tables": list({f["table_name"] for f in fields}),
        "total_tables": len({f["table_name"] for f in fields}),
        "total_fields": len(fields),
        "mapped_fields": sum(1 for f in fields if f.get("mapped")),
        "unmapped_fields": sum(1 for f in fields if not f.get("mapped")),
        "fields": fields,
        "scanned_at": now,
    }

    store._client.execute(
        "UPDATE discovery_scan_tasks SET status='completed', result_data=%s, tables_count=%s, fields_count=%s, mapped_fields=%s, unmapped_fields=%s, new_found=%s, completed_at=CURRENT_TIMESTAMP WHERE task_id=%s",
        (json.dumps(result, ensure_ascii=False, default=str), result["total_tables"], result["total_fields"], result["mapped_fields"], result["unmapped_fields"], new_count, task_id),
    )

    return DiscoveryScanResponse(task_id=task_id, status="completed")


@router.post("/discovery/scan", response_model=DiscoveryScanResponse)
async def start_discovery_scan(req: DiscoveryScanRequest | None = None):
    """Start a discovery scan task. Accepts optional source_config from frontend."""
    task_id = str(uuid.uuid4())
    config = (req.source_config if req and req.source_config else None) or {}
    sample_limit = req.sample_limit if req else 10000
    store = _get_discovery_store()
    store.create_task(task_id, config, sample_limit)
    return DiscoveryScanResponse(task_id=task_id, status="started")


@router.get("/discovery/scan/{task_id}/status")
async def get_scan_status(task_id: str):
    """SSE stream: runs discovery scan and yields progress events per table, then a done event."""
    store = _get_discovery_store()
    task = store.get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail=f"Scan task '{task_id}' not found")

    source_config = task.get("source_config") or {}
    if isinstance(source_config, str):
        source_config = json.loads(source_config)
    sample_limit = task.get("sample_limit", 10000)

    # 将 sample_limit 注入 source_config 传递给 discovery 服务
    if source_config.get("sqlserver"):
        source_config["sqlserver"]["sample_limit"] = sample_limit
    else:
        source_config["sample_limit"] = sample_limit

    async def event_generator():
        try:
            result = await asyncio.to_thread(_run_discovery_sync, source_config, store)
            tables = result.get("tables", [])
            total_fields = result.get("total_fields", 0)
            mapped_fields = result.get("mapped_fields", 0)
            unmapped_fields = result.get("unmapped_fields", 0)
            table_statuses = result.get("table_statuses", [])

            # 获取历史扫描中出现过的字段集合（排除本次任务，防止自己跟自己比）
            previously_seen = store.get_previously_scanned_fields(exclude_task_id=task_id)
            total_new = 0
            cached_count = 0
            scanned_count = 0

            # 预分组 table → fields（O(F) 一次），避免循环内重复全量过滤
            fields_by_table: dict[str, list[dict]] = {}
            for _f in result.get("fields", []):
                fields_by_table.setdefault(_f.get("table_name", ""), []).append(_f)

            for ts in table_statuses:
                table = ts["table"]
                table_fields = fields_by_table.get(table, [])
                # 新增字段 = 当前字段中，历史上从未出现过的「表名:字段名」
                new_count = 0
                for f in table_fields:
                    key = f"{f.get('table_name', '')}:{f.get('field_name', '')}".lower()
                    if key not in previously_seen:
                        new_count += 1
                total_new += new_count
                if ts.get("cached"):
                    cached_count += 1
                else:
                    scanned_count += 1
                yield f"event: progress\ndata: {json.dumps({'table': table, 'status': ts['status'], 'fields': len(table_fields), 'new': new_count, 'cached': ts.get('cached', False)})}\n\n"
                await asyncio.sleep(0.05)

            store.update_task_result(
                task_id, result,
                tables_count=len(tables),
                fields_count=total_fields,
                mapped_fields=mapped_fields,
                unmapped_fields=unmapped_fields,
                new_found=total_new,
            )

            # 扫描完成后批量刷新已映射指标的质量分（替代 GET 端点里的读时写入）
            try:
                _refresh_quality_scores_from_scan(result.get("fields", []))
            except Exception:
                logger.warning("扫描后刷新质量分失败", exc_info=True)

            yield f"event: done\ndata: {json.dumps({'status': 'completed', 'tables_scanned': len(tables), 'total_fields': total_fields, 'mapped_fields': mapped_fields, 'unmapped_fields': unmapped_fields, 'new_mappings': total_new})}\n\n"
        except Exception as exc:
            logger.error("Discovery scan failed: %s", exc, exc_info=True)
            store.update_task_error(task_id, str(exc))
            yield f"event: error\ndata: {json.dumps({'message': str(exc)})}\n\n"

    return StreamingResponse(event_generator(), media_type="text/event-stream")


@router.get("/discovery/results", response_model=DiscoveryResultsResponse)
def get_discovery_results():
    """Return the latest scan results from PostgreSQL, or empty if none cached."""
    store = _get_discovery_store()
    latest_result = store.get_latest_result()

    if latest_result is None:
        latest_result = {"tables": [], "total_tables": 0, "total_fields": 0, "mapped_fields": 0, "unmapped_fields": 0, "fields": []}

    fields: list[DiscoveryResultItem] = []
    tables_seen: set[str] = set()
    mapped_count = 0
    unmapped_count = 0

    # 从语义层 registry 实时构建已映射字段集合
    mapped_fields_set: set[str] = set()
    usage_map: dict[str, int] = {}
    try:
        reg = get_registry()
        for m in reg._store.list_metrics():
            if m.source_field:
                s = m.source_field.lower().strip()
                mapped_fields_set.add(s)
                usage_map[m.source_field] = m.usage_count
    except Exception:
        logger.warning("get_discovery_results: 构建 mapped_fields_set 失败，降级为全未映射", exc_info=True)

    # 批量预取全部字段释义，避免循环内逐条查询（N+1 → 1）
    try:
        desc_map = store.get_all_field_descriptions()
    except Exception:
        logger.warning("get_discovery_results: 批量加载字段释义失败，降级为空表", exc_info=True)
        desc_map = {}

    for f in latest_result.get("fields", []):
        tables_seen.add(f.get("table_name", ""))
        fn = (f.get("field_name") or "").lower().strip()
        tn = (f.get("table_name") or "").lower().strip()
        full = f"{tn}.{fn}"
        is_mapped = full in mapped_fields_set
        if is_mapped:
            mapped_count += 1
        else:
            unmapped_count += 1

        # 合并 Excel 导入的字段释义（从预取的 dict 查询，O(1)）
        table = f.get("table_name", "")
        field_name = f.get("field_name", "")
        description = f.get("description")
        is_primary_key = False
        remark = None
        desc_entry = desc_map.get(f"{table}:{field_name}".lower())
        if desc_entry:
            description = desc_entry.get("description") or description
            is_primary_key = desc_entry.get("is_primary_key", False)
            remark = desc_entry.get("remark")

        # 将合并后的描述写回 field dict，确保价值分计算使用最新描述
        f["description"] = description

        fields.append(DiscoveryResultItem(
            field_name=f.get("field_name", ""),
            table_name=f.get("table_name", ""),
            description=description,
            data_type=f.get("data_type", ""),
            non_null_rate=f.get("non_null_rate") or 0.0,
            non_null_row_count=f.get("non_null_row_count") or 0,
            distinct_count=f.get("distinct_count"),
            sample_value=f.get("sample_value"),
            sample_values=f.get("sample_values"),
            sample_stats=DiscoverySampleStats(**f.get("sample_stats")) if f.get("sample_stats") else None,
            is_dictionary=f.get("is_dictionary", False),
            last_updated=f.get("last_updated"),
            suggested_object=f.get("suggested_object"),
            mapped=is_mapped,
            table_schema=f.get("table_schema"),
            is_nullable=f.get("is_nullable"),
            is_primary_key=is_primary_key,
            remark=remark,
            quality_score=calc_field_quality(f),
            value_score=calc_field_value_score(f, usage_count=usage_map.get(f.get("field_name", ""), 0)),
        ))

    return DiscoveryResultsResponse(
        tables_count=len(tables_seen),
        fields_count=mapped_count + unmapped_count,
        mapped_fields=mapped_count,
        unmapped_fields=unmapped_count,
        fields=fields,
    )


@router.get("/discovery/history", response_model=list[DiscoveryHistoryItem])
def get_discovery_history():
    """Return discovery scan history from PostgreSQL."""
    store = _get_discovery_store()
    return [DiscoveryHistoryItem(**h) for h in store.get_scan_history()]


# ── 字段中文释义存储（Excel 导入，5列格式） ──────────────────────────


def _get_scanned_table_names() -> set[str]:
    """从 discovery 持久化存储中提取已扫描的原始表名集合。"""
    store = _get_discovery_store()
    return store.get_scanned_table_names()


@router.get("/field-descriptions/template")
def download_field_description_template():
    """下载字段中文释义 Excel 模版（5列：表名、是否主键、字段名、字段描述、备注）。"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "字段中文释义"

    headers = ["表名", "是否主键", "字段名", "字段描述", "备注"]
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 示例数据
    example_rows = [
        ["CASE_HISTORY_CONFIG", "是", "XML_CODE", "节点名称", ""],
        ["CASE_HISTORY_CONFIG", "是", "USE_TYPE", "适用病案类型", "非空；0：全部；1：西医；2：中医"],
        ["yb_settlement", "", "SET_NO", "结算流水号", ""],
    ]
    example_font = Font(name="微软雅黑", size=10, color="666666")
    for row_idx, row_data in enumerate(example_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = example_font
            cell.border = thin_border

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 35

    # 填写说明 sheet
    ws2 = wb.create_sheet("填写说明")
    instructions = [
        ["填写说明"],
        [""],
        ["1. 请勿修改表头（第一行）"],
        ["2. 表名：数据库中的表名（必填），必须先执行扫描才能导入"],
        ["3. 是否主键：是/否（可选），标记该字段是否为主键"],
        ["4. 字段名：数据库中的字段名（必填）"],
        ["5. 字段描述：字段的中文释义（必填）"],
        ["6. 备注：补充说明（可选），如值域范围、编码含义等"],
        ["7. 示例数据仅供格式参考，上传前请删除或替换为实际数据"],
        ["8. 导入前请先执行扫描，否则 Excel 中未扫描到的表名将被报错拒绝"],
    ]
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="微软雅黑", size=10)
    ws2.column_dimensions["A"].width = 80

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=field_description_template.xlsx"},
    )


@router.post("/field-descriptions/import")
def import_field_descriptions(file: UploadFile):
    """上传字段中文释义 Excel（5列），批量导入。

    表头：表名、是否主键、字段名、字段描述、备注（支持中英文列名，按位置回退）。
    校验：Excel 中的表名必须已存在于最近一次扫描结果中，否则报错并列出未知表名。
    """
    from io import BytesIO

    if not file.filename:
        raise HTTPException(status_code=400, detail="未选择文件")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("xlsx", "xls"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 格式的 Excel 文件")

    try:
        import openpyxl
        content = BytesIO(file.file.read())
        wb = openpyxl.load_workbook(content, read_only=True)
        ws = wb.active
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"无法读取 Excel 文件: {exc}") from exc

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 1:
        raise HTTPException(status_code=400, detail="Excel 文件为空")

    # 获取已扫描表名用于校验
    scanned_tables = _get_scanned_table_names()
    if not scanned_tables:
        raise HTTPException(
            status_code=400,
            detail="尚未执行过扫描，请先在页面点击「重新扫描」后再导入 Excel",
        )

    # 解析表头：名称匹配 → 位置回退
    header_row = rows[0]
    col_map: dict[str, int] = {}
    for idx, val in enumerate(header_row):
        if val and isinstance(val, str):
            key = val.strip().lower()
            if key in ("tablename", "table_name", "table"):
                col_map["table_name"] = idx
            elif key in ("isprimarykey", "is_primary_key", "pk"):
                col_map["is_primary_key"] = idx
            elif key in ("fieldname", "field_name", "field"):
                col_map["field_name"] = idx
            elif key in ("description", "desc", "comment"):
                col_map["description"] = idx
            elif key in ("remark", "remarks", "note"):
                col_map["remark"] = idx

    # 位置回退
    if "table_name" not in col_map and len(header_row) > 0:
        col_map["table_name"] = 0
    if "is_primary_key" not in col_map and len(header_row) > 1:
        col_map["is_primary_key"] = 1
    if "field_name" not in col_map and len(header_row) > 2:
        col_map["field_name"] = 2
    if "description" not in col_map and len(header_row) > 3:
        col_map["description"] = 3
    if "remark" not in col_map and len(header_row) > 4:
        col_map["remark"] = 4

    required = ["table_name", "field_name", "description"]
    missing = [c for c in required if c not in col_map]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Excel 表头必须包含：表名、字段名、字段描述（当前缺少：{', '.join(missing)}）",
        )

    # 第一遍：收集 Excel 中的所有表名
    excel_tables: set[str] = set()
    for row in rows[1:]:
        if not row:
            continue
        table_val = row[col_map["table_name"]]
        if table_val:
            excel_tables.add(str(table_val).strip().lower())

    scanned_lowered = {t.lower() for t in scanned_tables}
    unknown_tables = sorted(excel_tables - scanned_lowered)

    # 第二遍：解析数据行（跳过未知表，不中断导入）
    imported = 0
    skipped = 0
    skipped_rows: list[str] = []
    for row in rows[1:]:
        if not row:
            continue

        def _cell(idx: int) -> str:
            return str(row[idx]).strip() if len(row) > idx and row[idx] else ""

        table = _cell(col_map["table_name"])
        field = _cell(col_map["field_name"])
        desc = _cell(col_map["description"])

        if not table or not field or not desc:
            continue

        # 跳过未知表，记录但不中断
        if table.lower() not in scanned_lowered:
            skipped += 1
            if table not in skipped_rows:
                skipped_rows.append(table)
            continue

        pk_raw = _cell(col_map["is_primary_key"]).lower()
        is_pk = pk_raw in ("是", "yes", "true", "1", "y")
        remark = _cell(col_map["remark"]) or None

        store = _get_discovery_store()
        store.save_field_description(table, field, desc, is_pk, remark)
        imported += 1

    wb.close()

    msg = f"成功导入 {imported} 条字段释义"
    if skipped > 0:
        msg += f"，跳过 {skipped} 条（{len(skipped_rows)} 个未扫描的表：{'、'.join(skipped_rows[:5])}"
        if len(skipped_rows) > 5:
            msg += f" 等 {len(skipped_rows)} 个"
        msg += "），请先执行扫描后再导入"

    return {
        "status": "ok",
        "imported_count": imported,
        "skipped_count": skipped,
        "skipped_tables": skipped_rows,
        "message": msg,
    }


@router.get("/field-descriptions/count")
def get_description_store_count():
    """返回当前已导入的字段释义数量。"""
    store = _get_discovery_store()
    return {"count": store.get_field_descriptions_count()}


class FieldDescriptionUpdateRequest(BaseModel):
    table: str
    field: str
    description: str
    remark: str | None = None


@router.put("/field-descriptions")
def update_field_description(req: FieldDescriptionUpdateRequest):
    """更新单个字段的描述和备注。"""
    store = _get_discovery_store()
    store.save_field_description(
        table_name=req.table,
        field_name=req.field,
        description=req.description,
        remark=req.remark,
    )
    return {"status": "ok", "message": f"已更新 {req.table}.{req.field} 的字段释义"}
