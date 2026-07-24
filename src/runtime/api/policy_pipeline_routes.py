"""
政策知识管线 API

端点:
  GET  /policy-pipeline/summary                — 管线概览统计
  GET  /policy-pipeline/documents              — 原文列表
  POST /policy-pipeline/documents              — 创建原文
  GET  /policy-pipeline/documents/template     — 下载 Excel 模板
  POST /policy-pipeline/documents/upload       — 批量上传 Excel
  GET  /policy-pipeline/documents/{id}         — 原文详情
  PUT  /policy-pipeline/documents/{id}         — 更新原文
  DELETE /policy-pipeline/documents/{id}       — 删除原文
  POST /policy-pipeline/documents/{id}/extract — 触发提取
  GET  /policy-pipeline/extractions            — 提取结果列表
  GET  /policy-pipeline/extractions/{id}       — 提取结果详情
  PUT  /policy-pipeline/extractions/{id}       — 更新提取结果
  DELETE /policy-pipeline/extractions/{id}     — 删除提取结果
  POST /policy-pipeline/extractions/{id}/publish — 发布到 Milvus
  GET  /policy-pipeline/rules/{rule_id}/lineage  — 溯源信息
"""
from __future__ import annotations

import io
import logging

from fastapi import APIRouter, HTTPException, Query, Request, UploadFile
from fastapi.responses import StreamingResponse

from src.knowledge_extension.rule_explanation.pipeline_store import PipelineStore
from src.knowledge_extension.rule_explanation.pipeline_orchestrator import PipelineOrchestrator
from src.shared.schemas.responses import error_detail
from src.data_platform.storage.postgresql.policy_meta_store import PolicyMetaStore
from pydantic import BaseModel

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/api/v1/medical-insurance-ai-agent/policy-pipeline",
    tags=["policy-pipeline"],
)

_store: PipelineStore | None = None
_orchestrator: PipelineOrchestrator | None = None


def _get_store() -> PipelineStore:
    global _store
    if _store is None:
        _store = PipelineStore()
    return _store


def _get_orchestrator() -> PipelineOrchestrator:
    global _orchestrator
    if _orchestrator is None:
        _orchestrator = PipelineOrchestrator(_get_store())
    return _orchestrator


# ═══════════════ Summary ═══════════════

@router.get("/summary")
def get_summary():
    return _get_store().get_summary()


# ═══════════════ Documents — list / create ═══════════════

@router.get("/documents")
def list_documents(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    status: str = Query(""),
    keyword: str = Query(""),
):
    return _get_store().list_documents(page=page, page_size=page_size, status=status, keyword=keyword)


@router.post("/documents")
async def create_document(request: Request):
    body = await request.json()
    if not body.get("title") or not body.get("content_text"):
        raise HTTPException(status_code=400, detail=error_detail("INVALID_INPUT", "title 和 content_text 为必填项", {}))
    return _get_store().create_document(body)


# ═══════════════ Documents — template / upload (before {doc_id}!) ═══════════════

DOC_TEMPLATE_HEADERS = [
    "标题", "主题分类", "发文机构", "发文字号", "文件来源",
    "发布日期", "实施日期", "废止日期", "成文日期", "有效性",
    "详情页URL", "正文内容",
]

DOC_TEMPLATE_EXAMPLE = [
    "北京市人力资源和社会保障局印发《北京市城乡居民基本医疗保险办法实施细则》的通知",
    "财政、金融、审计/保险",
    "北京市人力资源和社会保障局",
    "京人社农合发 [2017] 250号",
    "",
    "2017-12-13",
    "",
    "",
    "2017-11-30",
    "现行有效",
    "https://ybj.beijing.gov.cn/zwgk/2024zcwj/202406/t20240612_3710806.html",
    "第一条 参保人员在定点医疗机构发生的住院医疗费用，起付标准为1300元。\n第二条 起付标准以上至3万元的部分，统筹基金支付85%，职工支付15%。\n第三条 超过3万元至4万元的部分，统筹基金支付90%，职工支付10%。\n第四条 超过4万元的部分，统筹基金支付95%，职工支付5%。",
]

DOC_TEMPLATE_WIDTHS = [40, 18, 20, 22, 16, 14, 14, 14, 14, 12, 35, 60]


@router.get("/documents/template")
def download_document_template():
    """下载政策原文 Excel 导入模板"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "政策原文批量导入"

    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(DOC_TEMPLATE_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    example_font = Font(name="微软雅黑", size=10, color="666666")
    for col_idx, value in enumerate(DOC_TEMPLATE_EXAMPLE, 1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.font = example_font
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for idx, width in enumerate(DOC_TEMPLATE_WIDTHS, 1):
        col_letter = chr(64 + idx) if idx <= 26 else "A"
        ws.column_dimensions[col_letter].width = width

    # 填写说明 sheet
    ws2 = wb.create_sheet("填写说明")
    instructions = [
        ["政策原文批量导入 — 填写说明"],
        [""],
        ["1. 请勿修改表头（第一行），否则会导致导入失败"],
        ["2. 标题和正文内容为必填项，其他字段可选"],
        ["3. 日期格式：YYYY-MM-DD（如 2017-12-13）"],
        ["4. 有效性：现行有效 / 有效 / 是 → valid；已废止 / 废止 → abolished；未知 → unknown"],
        ["5. 示例数据仅供格式参考，上传前请删除或替换为实际数据"],
        ["6. 正文内容中的换行符会自动保留"],
        [""],
        ["列说明："],
        ["- 标题 *：政策文件的完整标题"],
        ["- 主题分类：如 财政金融审计/保险、医保待遇、基金管理"],
        ["- 发文机构：如 北京市人力资源和社会保障局"],
        ["- 发文字号：如 京人社农合发 [2017] 250号"],
        ["- 文件来源：政策来源方（可选）"],
        ["- 发布日期：政策公开发布日期"],
        ["- 实施日期：政策正式生效日期"],
        ["- 废止日期：如已废止则填写（可选）"],
        ["- 成文日期：文件签批日期"],
        ["- 有效性：现行有效 / 有效 / 已废止 等"],
        ["- 详情页URL：爬虫来源的原始网页地址（可选）"],
        ["- 正文内容 *：政策的完整正文"],
    ]
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="微软雅黑", size=10, bold=(row_idx == 1))
    ws2.column_dimensions["A"].width = 80

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=policy_document_template.xlsx"},
    )


# 表头→字段映射（支持中英文）
HEADER_FIELD_MAP: dict[str, str] = {
    "标题": "title", "title": "title",
    "主题分类": "category", "category": "category",
    "发文机构": "issuing_agency", "issuing_agency": "issuing_agency",
    "发文字号": "document_number", "document_number": "document_number",
    "文件来源": "file_source", "file_source": "file_source",
    "发布日期": "publish_date", "publish_date": "publish_date",
    "实施日期": "effective_date", "effective_date": "effective_date",
    "废止日期": "abolition_date", "abolition_date": "abolition_date",
    "成文日期": "document_date", "document_date": "document_date",
    "有效性": "validity", "validity": "validity",
    "详情页url": "source_url", "详情页URL": "source_url", "source_url": "source_url",
    "正文内容": "content_text", "content_text": "content_text",
}

POS_FALLBACK = ["title", "category", "issuing_agency", "document_number", "file_source",
                "publish_date", "effective_date", "abolition_date",
                "document_date", "validity", "source_url", "content_text"]

VALIDITY_VALUES = {"现行有效": "valid", "有效": "valid", "是": "valid", "已废止": "abolished", "废止": "abolished", "未知": "unknown"}


@router.post("/documents/upload")
async def upload_documents(file: UploadFile):
    """批量上传政策原文 Excel"""
    from openpyxl import load_workbook

    if not file.filename:
        raise HTTPException(status_code=400, detail=error_detail("NO_FILE", "未选择文件", {}))
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("xlsx", "xls"):
        raise HTTPException(status_code=400, detail=error_detail("BAD_FORMAT", "仅支持 .xlsx/.xls", {}))

    try:
        content = io.BytesIO(await file.read())
        wb = load_workbook(content, read_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=error_detail("READ_ERROR", str(e), {}))

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail=error_detail("EMPTY", "Excel 至少需要表头+1行数据", {}))

    header_row = rows[0]
    col_field: dict[int, str] = {}
    for idx, val in enumerate(header_row):
        if val and isinstance(val, str):
            key = val.strip()
            field = HEADER_FIELD_MAP.get(key, HEADER_FIELD_MAP.get(key.lower(), ""))
            if field:
                col_field[idx] = field
    for idx, field in enumerate(POS_FALLBACK):
        if idx not in col_field and idx < len(header_row):
            col_field[idx] = field

    if "title" not in col_field.values() or "content_text" not in col_field.values():
        raise HTTPException(status_code=400, detail=error_detail("BAD_HEADER", "表头必须包含「标题」和「政策原文内容」", {}))

    store = _get_store()
    created = 0
    skipped = 0
    errors: list[str] = []

    for row_idx, row in enumerate(rows[1:], 2):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        data: dict[str, str] = {}
        for col_idx, field in col_field.items():
            val = row[col_idx] if col_idx < len(row) else None
            data[field] = str(val).strip() if val is not None else ""

        if not data.get("title") or not data.get("content_text"):
            skipped += 1
            errors.append(f"第{row_idx}行：标题或内容为空")
            continue
        # 值域标准化
        v = data.get("validity", "").strip()
        if v in VALIDITY_VALUES:
            data["validity"] = VALIDITY_VALUES[v]
        if not data.get("source_type"):
            data["source_type"] = "manual"

        try:
            store.create_document(data)
            created += 1
        except Exception as e:
            skipped += 1
            errors.append(f"第{row_idx}行：{e}")

    return {"created": created, "skipped": skipped, "total": created + skipped, "errors": errors[:20]}


# ═══════════════ Documents — detail / update / delete / extract ═══════════════

@router.get("/documents/{doc_id}")
def get_document(doc_id: str):
    doc = _get_store().get_document(doc_id)
    if not doc:
        raise HTTPException(status_code=404, detail=error_detail("NOT_FOUND", "文档不存在", {"doc_id": doc_id}))
    doc["rule_ids"] = _get_store().get_rules_by_doc(doc_id)
    return doc


@router.put("/documents/{doc_id}")
async def update_document(doc_id: str, request: Request):
    body = await request.json()
    doc = _get_store().update_document(doc_id, body)
    if not doc:
        raise HTTPException(status_code=404, detail=error_detail("NOT_FOUND", "文档不存在", {"doc_id": doc_id}))
    return doc


@router.delete("/documents/{doc_id}")
def delete_document(doc_id: str):
    if not _get_store().delete_document(doc_id):
        raise HTTPException(status_code=404, detail=error_detail("NOT_FOUND", "文档不存在", {"doc_id": doc_id}))
    return {"deleted": True, "doc_id": doc_id}


@router.post("/documents/{doc_id}/extract")
def trigger_extraction(doc_id: str):
    result = _get_orchestrator().run_extraction(doc_id)
    if not result.get("success"):
        raise HTTPException(status_code=500, detail=error_detail("EXTRACTION_FAILED", result.get("error", ""), {"doc_id": doc_id}))
    return result


# ═══════════════ Extractions ═══════════════

@router.get("/extractions")
def list_extractions(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    doc_id: str = Query(""),
    status: str = Query(""),
):
    return _get_store().list_extractions(page=page, page_size=page_size, doc_id=doc_id, status=status)


@router.get("/extractions/{extraction_id}")
def get_extraction(extraction_id: str):
    ext = _get_store().get_extraction(extraction_id)
    if not ext:
        raise HTTPException(status_code=404, detail=error_detail("NOT_FOUND", "提取记录不存在", {"extraction_id": extraction_id}))
    return ext


@router.put("/extractions/{extraction_id}")
async def update_extraction(extraction_id: str, request: Request):
    body = await request.json()
    ext = _get_store().update_extraction(extraction_id, body)
    if not ext:
        raise HTTPException(status_code=404, detail=error_detail("NOT_FOUND", "提取记录不存在", {"extraction_id": extraction_id}))
    return ext


@router.delete("/extractions/{extraction_id}")
def delete_extraction(extraction_id: str):
    if not _get_store().delete_extraction(extraction_id):
        raise HTTPException(status_code=404, detail=error_detail("NOT_FOUND", "提取记录不存在", {"extraction_id": extraction_id}))
    return {"deleted": True, "extraction_id": extraction_id}


@router.post("/extractions/{extraction_id}/publish")
def publish_extraction(extraction_id: str):
    result = _get_orchestrator().publish_extraction(extraction_id)
    if not result.get("success"):
        raise HTTPException(status_code=500, detail=error_detail("PUBLISH_FAILED", result.get("error", ""), {"extraction_id": extraction_id}))
    return result


@router.post("/extractions/{extraction_id}/publish-v2")
def publish_extraction_v2(extraction_id: str):
    """发布到新 collection（policy_facts + policy_rules_v2，P3 新通路）。

    与 /publish（写旧 policy_rules）并存；P10 切换后此端点成为主入口。
    """
    result = _get_orchestrator().publish_to_new_collections(extraction_id)
    if not result.get("success"):
        raise HTTPException(status_code=500, detail=error_detail(
            "PUBLISH_FAILED", result.get("error", ""), {"extraction_id": extraction_id}))
    return result


# ═══════════════ Rules — lineage ═══════════════

@router.post("/rules/search")
async def search_rules(request: Request):
    """政策规则混合检索（P6，基于 policy_rules_v2 + 业务库）。

    body: {mode, query?, filters?, target, metric_codes?, context?, top_k?}
    - mode: precise|semantic|hybrid（target=policy/both 时生效）
    - target: policy(默认)|database|both
      - policy: 查政策规则（三模式）
      - database: 查业务数据（经 source_field 映射查 SQLServer，需 metric_codes+context）
      - both: 政策规则 + 业务数据
    """
    from src.knowledge_extension.rule_explanation.rules_search_service import RulesSearchService
    body = await request.json()
    mode = body.get("mode", "precise")
    target = body.get("target", "policy")
    top_k = int(body.get("top_k", 20))
    svc = RulesSearchService()

    groups: list = []
    if target in ("policy", "both"):
        if mode == "precise":
            groups = svc.search_precise(body.get("filters", {}), top_k=top_k)
        elif mode == "semantic":
            groups = svc.search_semantic(body["query"], top_k=top_k)
        elif mode == "hybrid":
            groups = svc.search_hybrid(body["query"], body.get("filters", {}), top_k=top_k)
        else:
            raise HTTPException(status_code=400, detail=error_detail(
                "INVALID_MODE", f"mode 必须是 precise/semantic/hybrid，实际={mode}", {}))

    database_values: dict = {}
    if target in ("database", "both"):
        metric_codes = body.get("metric_codes", [])
        if not metric_codes:
            raise HTTPException(status_code=400, detail=error_detail(
                "NO_METRICS", "target=database/both 必须提供 metric_codes", {}))
        database_values = svc.search_database(metric_codes, body.get("context", {}))

    return {
        "mode": mode, "target": target,
        "groups": groups, "total_groups": len(groups),
        "database_values": database_values,
    }


@router.get("/rules/{rule_id}/lineage")
def get_rule_lineage(rule_id: str):
    return {"rule_id": rule_id, "lineages": _get_store().get_lineages_by_rule(rule_id)}


# ═══════════════ Rules — unpublished (from extractions) ═══════════════

@router.get("/rules/unpublished")
def list_unpublished_rules(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    doc_id: str = Query(""),
    rule_type: str = Query(""),
):
    """获取未发布的规则（从提取记录中展开）。状态 draft/reviewed 的提取记录中的规则。"""
    store = _get_store()
    ext_list = store.list_extractions(
        page=1, page_size=10000,
        doc_id=doc_id, status="",
    )
    items = ext_list.get("items", [])

    # 展开每条提取记录中的规则
    flat_rules: list[dict] = []
    for ext in items:
        if ext["status"] in ("published", "rejected"):
            continue
        fields = ext.get("extracted_fields", {})
        if isinstance(fields, str):
            try:
                import json as _json
                fields = _json.loads(fields)
            except Exception:
                fields = {}
        rules = fields.get("rules", [])
        if not rules:
            # 旧格式兼容：提取字段本身作为单条规则
            if fields.get("rule_type"):
                rules = [fields]
        for rule in rules:
            rule["extraction_id"] = ext["extraction_id"]
            rule["doc_id"] = ext["doc_id"]
            rule["doc_title"] = ext.get("doc_title", "")
            rule["segment_status"] = ext["status"]
            rule["fact_text"] = fields.get("fact_text", ext.get("source_text", ""))
            if rule_type and rule.get("rule_type") != rule_type:
                continue
            flat_rules.append(rule)

    total = len(flat_rules)
    start = (page - 1) * page_size
    end = start + page_size
    page_items = flat_rules[start:end]

    return {"items": page_items, "total": total, "page": page, "page_size": page_size}


# ════════════════════ schema-update 重新结构化任务（P5.4，§7.3/§7.4）═════════════════════

_meta_store: PolicyMetaStore | None = None


def _get_meta_store() -> PolicyMetaStore:
    """PolicyMetaStore 单例（测试可 monkeypatch 注入内存假 store）。"""
    global _meta_store
    if _meta_store is None:
        _meta_store = PolicyMetaStore()
    return _meta_store


class SchemaUpdatePublishRequest(BaseModel):
    metric_code: str
    strategy: str  # incremental | full | soft_delete
    change_type: str = "modify"  # add | modify | remove
    schema_version: int = 1
    affected_docs: list[str] = []  # 受影响文档（调用方提供，绕开 metric_code 反查缺口）
    new_values: dict | None = None  # incremental/full 新值（LLM 提取后续接）


def _build_schema_executor():
    """构造 schema 更新执行器：reader=query_rules_by_doc，writer=upsert_rules（真实 Milvus）。

    测试可通过 monkeypatch 替换为 fake executor。真实执行需 Milvus 可用。
    """
    from src.knowledge_extension.rule_explanation.schema_update_executor import SchemaUpdateExecutor
    from src.knowledge_extension.rule_explanation.policy_retrieval.policy_rules_schema_v2 import (
        query_rules_by_doc, upsert_rules, POLICY_RULES_V2_COLLECTION,
    )
    from pymilvus import Collection

    def _writer(entities):
        col = Collection(POLICY_RULES_V2_COLLECTION)
        col.load()
        return upsert_rules(col, entities)

    return SchemaUpdateExecutor(
        reader=lambda doc_id: query_rules_by_doc(doc_id),
        writer=_writer,
    )


@router.post("/schema-update/publish")
def publish_schema_update(req: SchemaUpdatePublishRequest):
    """触发 schema 更新执行（§7.3 read-modify-write）。

    - 无 affected_docs：仅创建 task（status=pending，兼容旧调用）。
    - 有 affected_docs：create task → 逐 doc evolve（read→modify→write）→ 标记 done。
      incremental/full 需 new_values（LLM 提取后续接，当前由调用方提供）。
    """
    task = _get_meta_store().create_task(
        metric_code=req.metric_code, change_type=req.change_type,
        strategy=req.strategy, schema_version=req.schema_version,
    )
    task_id = task["task_id"]

    if not req.affected_docs:
        return {"task_id": task_id, "status": "pending"}

    executor = _build_schema_executor()
    meta = _get_meta_store()
    grand_processed = 0
    grand_total = 0
    try:
        for doc_id in req.affected_docs:
            summary = executor.evolve(
                doc_id, req.strategy, new_values=req.new_values,
                schema_version=req.schema_version,
                on_progress=lambda p, t: meta.update_task_progress(
                    task_id, grand_processed + p, grand_total + t, status="running"),
            )
            grand_processed += summary["processed"]
            grand_total += summary["total"]
        meta.update_task_progress(task_id, grand_processed, grand_total, status="done")
        return {
            "task_id": task_id, "status": "done",
            "summary": {"processed": grand_processed, "total": grand_total,
                        "docs": len(req.affected_docs)},
        }
    except Exception as exc:
        meta.fail_task(task_id, str(exc))
        raise HTTPException(status_code=500, detail=f"schema 更新执行失败: {exc}")


@router.get("/schema-update/tasks/{task_id}")
def get_schema_update_task(task_id: str):
    task = _get_meta_store().get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail=f"任务 '{task_id}' 不存在")
    return task


@router.get("/schema-update/tasks")
def list_schema_update_tasks(
    status: str = Query(""), metric_code: str = Query("")
):
    return _get_meta_store().list_tasks(status=status, metric_code=metric_code)
